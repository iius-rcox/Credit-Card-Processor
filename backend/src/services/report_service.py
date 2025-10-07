"""
ReportService - Handles report generation in Excel and CSV formats.

This module generates reconciliation reports with transaction and receipt data.
"""

import csv
import io
from typing import BinaryIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..models.session import Session
from ..repositories.employee_repository import EmployeeRepository
from ..repositories.match_result_repository import MatchResultRepository
from ..repositories.receipt_repository import ReceiptRepository
from ..repositories.session_repository import SessionRepository
from ..repositories.transaction_repository import TransactionRepository


class ReportService:
    """
    Service for generating reconciliation reports.

    Provides methods for creating Excel (XLSX) and CSV reports with
    transaction and receipt matching data.
    """

    def __init__(
        self,
        session_repo: SessionRepository,
        employee_repo: EmployeeRepository,
        transaction_repo: TransactionRepository,
        receipt_repo: ReceiptRepository,
        match_result_repo: MatchResultRepository
    ):
        """
        Initialize report service.

        Args:
            session_repo: SessionRepository instance
            employee_repo: EmployeeRepository instance
            transaction_repo: TransactionRepository instance
            receipt_repo: ReceiptRepository instance
            match_result_repo: MatchResultRepository instance
        """
        self.session_repo = session_repo
        self.employee_repo = employee_repo
        self.transaction_repo = transaction_repo
        self.receipt_repo = receipt_repo
        self.match_result_repo = match_result_repo

    async def generate_excel_report(self, session_id: UUID) -> bytes:
        """
        Generate Excel (XLSX) report for a session.

        Args:
            session_id: UUID of the session

        Returns:
            Bytes of Excel file

        Note:
            Creates workbook with 3 sheets:
            1. Summary - Session metadata and statistics
            2. Transactions - All transactions with match status
            3. Receipts - All receipts with match references
        """
        # Get session data
        session = await self.session_repo.get_session_by_id(session_id)
        if not session:
            raise ValueError(f"Session {session_id} not found")

        employees = await self.employee_repo.get_employees_by_session(session_id)
        transactions = await self.transaction_repo.get_transactions_by_session(session_id)
        receipts = await self.receipt_repo.get_receipts_by_session(session_id)
        matches = await self.match_result_repo.get_match_results_by_session(session_id)

        # Create workbook
        wb = Workbook()

        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(ws_summary, session, employees, transactions, receipts, matches)

        # Sheet 2: Transactions
        ws_trans = wb.create_sheet("Transactions")
        self._create_transactions_sheet(ws_trans, transactions, matches)

        # Sheet 3: Receipts
        ws_receipts = wb.create_sheet("Receipts")
        self._create_receipts_sheet(ws_receipts, receipts, matches)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_summary_sheet(self, ws, session, employees, transactions, receipts, matches):
        """Create summary sheet with session metadata and statistics."""
        # Title
        ws["A1"] = "Credit Card Reconciliation Report"
        ws["A1"].font = Font(size=16, bold=True)

        # Session info
        row = 3
        ws[f"A{row}"] = "Session ID:"
        ws[f"B{row}"] = str(session.id)
        row += 1
        ws[f"A{row}"] = "Created:"
        ws[f"B{row}"] = session.created_at.strftime("%Y-%m-%d %H:%M:%S")
        row += 1
        ws[f"A{row}"] = "Status:"
        ws[f"B{row}"] = session.status
        row += 1
        ws[f"A{row}"] = "Expires:"
        ws[f"B{row}"] = session.expires_at.strftime("%Y-%m-%d %H:%M:%S")

        # Statistics
        row += 2
        ws[f"A{row}"] = "Statistics"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        row += 1

        matched_count = len([m for m in matches if m.match_status == "matched"])
        unmatched_count = len([m for m in matches if m.match_status == "unmatched"])
        review_count = len([m for m in matches if m.match_status == "manual_review"])

        stats = [
            ("Employees", len(employees)),
            ("Transactions", len(transactions)),
            ("Receipts", len(receipts)),
            ("Matched", matched_count),
            ("Unmatched", unmatched_count),
            ("Needs Review", review_count),
        ]

        for label, value in stats:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

        # Employee summary
        if employees:
            row += 2
            ws[f"A{row}"] = "Employees"
            ws[f"A{row}"].font = Font(size=14, bold=True)
            row += 1

            # Headers
            headers = ["Employee Number", "Name", "Department", "Cost Center"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            row += 1

            # Data
            for emp in employees:
                ws.cell(row=row, column=1, value=emp.employee_number)
                ws.cell(row=row, column=2, value=emp.name)
                ws.cell(row=row, column=3, value=emp.department or "")
                ws.cell(row=row, column=4, value=emp.cost_center or "")
                row += 1

        # Auto-size columns
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_transactions_sheet(self, ws, transactions, matches):
        """Create transactions sheet with match status."""
        # Headers
        headers = [
            "Transaction ID", "Date", "Amount", "Currency", "Merchant",
            "Description", "Card Last 4", "Match Status", "Receipt ID",
            "Confidence", "Amount Diff", "Date Diff (days)"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data
        match_dict = {m.transaction_id: m for m in matches}

        for row, trans in enumerate(transactions, start=2):
            match = match_dict.get(trans.id)

            ws.cell(row=row, column=1, value=str(trans.id))
            ws.cell(row=row, column=2, value=trans.transaction_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=3, value=float(trans.amount))
            ws.cell(row=row, column=4, value=trans.currency)
            ws.cell(row=row, column=5, value=trans.merchant_name)
            ws.cell(row=row, column=6, value=trans.description or "")
            ws.cell(row=row, column=7, value=trans.card_last_four or "")

            if match:
                ws.cell(row=row, column=8, value=match.match_status)
                ws.cell(row=row, column=9, value=str(match.receipt_id) if match.receipt_id else "")
                ws.cell(row=row, column=10, value=float(match.confidence_score))
                ws.cell(row=row, column=11, value=float(match.amount_difference) if match.amount_difference else "")
                ws.cell(row=row, column=12, value=match.date_difference_days if match.date_difference_days is not None else "")

        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_receipts_sheet(self, ws, receipts, matches):
        """Create receipts sheet with match references."""
        # Headers
        headers = [
            "Receipt ID", "Date", "Amount", "Currency", "Vendor",
            "File Name", "OCR Confidence", "Processing Status",
            "Matched Transaction ID"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Create reverse match lookup (receipt_id -> transaction_id)
        receipt_matches = {m.receipt_id: m.transaction_id for m in matches if m.receipt_id}

        # Data
        for row, receipt in enumerate(receipts, start=2):
            ws.cell(row=row, column=1, value=str(receipt.id))
            ws.cell(row=row, column=2, value=receipt.receipt_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=3, value=float(receipt.amount))
            ws.cell(row=row, column=4, value=receipt.currency)
            ws.cell(row=row, column=5, value=receipt.vendor_name)
            ws.cell(row=row, column=6, value=receipt.file_name)
            ws.cell(row=row, column=7, value=float(receipt.ocr_confidence) if receipt.ocr_confidence else "")
            ws.cell(row=row, column=8, value=receipt.processing_status)

            # Add matched transaction ID if exists
            matched_trans_id = receipt_matches.get(receipt.id)
            ws.cell(row=row, column=9, value=str(matched_trans_id) if matched_trans_id else "")

        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    async def generate_csv_report(self, session_id: UUID) -> str:
        """
        Generate CSV report for a session.

        Args:
            session_id: UUID of the session

        Returns:
            CSV string

        Note:
            CSV format includes combined transaction and receipt data in one row:
            Transaction ID, Date, Amount, Merchant, Match Status,
            Receipt ID, Receipt Amount, Confidence
        """
        # Get session data
        session = await self.session_repo.get_session_by_id(session_id)
        if not session:
            raise ValueError(f"Session {session_id} not found")

        transactions = await self.transaction_repo.get_transactions_by_session(session_id)
        receipts = await self.receipt_repo.get_receipts_by_session(session_id)
        matches = await self.match_result_repo.get_match_results_by_session(session_id)

        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)

        # Headers
        writer.writerow([
            "Transaction ID", "Transaction Date", "Transaction Amount",
            "Merchant", "Match Status", "Receipt ID", "Receipt Date",
            "Receipt Amount", "Vendor", "Confidence", "Amount Difference",
            "Date Difference (days)"
        ])

        # Create lookups
        match_dict = {m.transaction_id: m for m in matches}
        receipt_dict = {r.id: r for r in receipts}

        # Data rows
        for trans in transactions:
            match = match_dict.get(trans.id)
            receipt = receipt_dict.get(match.receipt_id) if match and match.receipt_id else None

            row = [
                str(trans.id),
                trans.transaction_date.strftime("%Y-%m-%d"),
                f"{trans.amount:.2f}",
                trans.merchant_name,
                match.match_status if match else "unmatched",
                str(receipt.id) if receipt else "",
                receipt.receipt_date.strftime("%Y-%m-%d") if receipt else "",
                f"{receipt.amount:.2f}" if receipt else "",
                receipt.vendor_name if receipt else "",
                f"{match.confidence_score:.4f}" if match else "",
                f"{match.amount_difference:.2f}" if match and match.amount_difference else "",
                str(match.date_difference_days) if match and match.date_difference_days is not None else ""
            ]

            writer.writerow(row)

        output.seek(0)
        return output.getvalue()
