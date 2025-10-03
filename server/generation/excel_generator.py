"""
Excel report generator for incomplete expenses.

Generates Excel (.xlsx) files listing all expenses missing receipts or GL/project codes,
with conditional formatting on the Status column.
"""

import os
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from api.models import Session, Employee, ExpenseTransaction, ExcelReport
from processing.analyzer import get_incomplete_expenses


# Status column colors (from spec clarification)
COLOR_MISSING_BOTH = "FFC7CE"  # Red
COLOR_MISSING_RECEIPT = "FFEB9C"  # Yellow
COLOR_MISSING_GL_CODE = "FFD966"  # Orange
COLOR_HEADER = "D3D3D3"  # Light gray


def generate_excel_report(session: Session, output_dir: str = "server/reports") -> str:
    """
    Generate Excel report listing all incomplete expenses.

    Report Structure:
    - Columns: Employee ID, Employee Name, Card Number, Transaction Date,
               Transaction Amount, Transaction Name, Status
    - Header row: Bold, gray background
    - Status column conditional formatting:
      * "Missing Both" = red background
      * "Missing Receipt" = yellow background
      * "Missing GL Code" = orange background
    - Only includes expenses where status != "Complete"
    - Sorted by Employee ID, then Transaction Date

    Args:
        session: Session object with employees and expenses
        output_dir: Directory to save Excel file

    Returns:
        Absolute path to generated Excel file

    Example:
        >>> path = generate_excel_report(session)
        >>> print(f"Excel report saved to {path}")
    """
    # Create output directory if doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # Generate file path
    session_id_str = str(session.session_id)
    file_path = os.path.join(output_dir, f"{session_id_str}_incomplete_expenses.xlsx")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Incomplete Expenses"

    # Write header row
    headers = [
        "Employee ID",
        "Employee Name",
        "Card Number",
        "Transaction Date",
        "Transaction Amount",
        "Transaction Name",
        "Status",
    ]

    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    header_font = Font(bold=True)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Get incomplete expenses from all employees
    incomplete_expenses = get_incomplete_expenses(session)

    # Sort by employee ID, then transaction date
    incomplete_expenses.sort(key=lambda e: (e.employee_id, e.transaction_date))

    # Write data rows
    row_num = 2
    for expense in incomplete_expenses:
        # Find employee for this expense
        employee = next((e for e in session.employees if e.employee_id == expense.employee_id), None)

        if not employee:
            continue  # Skip if employee not found

        # Mask card number (show only last 4 digits)
        masked_card = f"**** **** **** {employee.card_number[-4:]}"

        # Write row
        row_data = [
            expense.employee_id,
            employee.name,
            masked_card,
            expense.transaction_date.isoformat(),
            float(expense.transaction_amount),
            expense.transaction_name,
            expense.status.value,
        ]

        ws.append(row_data)

        # Apply conditional formatting to Status column
        status_cell = ws.cell(row=row_num, column=7)  # Status is column 7

        if expense.status.value == "Missing Both":
            status_cell.fill = PatternFill(
                start_color=COLOR_MISSING_BOTH, end_color=COLOR_MISSING_BOTH, fill_type="solid"
            )
        elif expense.status.value == "Missing Receipt":
            status_cell.fill = PatternFill(
                start_color=COLOR_MISSING_RECEIPT,
                end_color=COLOR_MISSING_RECEIPT,
                fill_type="solid",
            )
        elif expense.status.value == "Missing GL Code":
            status_cell.fill = PatternFill(
                start_color=COLOR_MISSING_GL_CODE,
                end_color=COLOR_MISSING_GL_CODE,
                fill_type="solid",
            )

        row_num += 1

    # Auto-width columns
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0

        for cell in ws[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        # Set column width (add padding)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Freeze top row for scrolling
    ws.freeze_panes = "A2"

    # Save workbook
    wb.save(file_path)

    return os.path.abspath(file_path)


def create_excel_report_metadata(session: Session, file_path: str) -> ExcelReport:
    """
    Create ExcelReport metadata object.

    Args:
        session: Session object
        file_path: Path to generated Excel file

    Returns:
        ExcelReport object with metadata
    """
    # Count incomplete expenses
    incomplete_count = len(get_incomplete_expenses(session))

    return ExcelReport(
        session_id=session.session_id,
        file_path=file_path,
        row_count=incomplete_count,
    )
