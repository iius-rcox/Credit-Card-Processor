"""
Unit tests for Excel generator (T051).

Tests Excel report structure, formatting, and content filtering.
"""

import pytest
import os
from decimal import Decimal
from datetime import date
from openpyxl import load_workbook

from api.models import Session, Employee, ExpenseTransaction
from generation.excel_generator import generate_excel_report


class TestExcelGenerator:
    """Test Excel report generation."""

    def test_excel_file_structure_7_columns(self):
        """Verify Excel has 7 columns as specified."""
        # Create minimal session with one incomplete expense
        expense = ExpenseTransaction(
            employee_id="EMP001",
            transaction_date=date(2025, 9, 15),
            transaction_amount=Decimal("125.50"),
            transaction_name="Test",
            has_receipt=False,
            has_gl_code=False,
        )

        employee = Employee(
            employee_id="EMP001",
            name="Test User",
            card_number="1234567890123456",
            expenses=[expense],
        )

        session = Session(
            credit_card_pdf_path="test.pdf",
            expense_report_pdf_path="test2.pdf",
            employees=[employee],
        )

        # Generate Excel
        file_path = generate_excel_report(session, output_dir="server/tests/unit/temp")

        # Load and verify
        wb = load_workbook(file_path)
        ws = wb.active

        # Check column count
        assert ws.max_column == 7, "Excel should have 7 columns"

        # Verify header row
        expected_headers = [
            "Employee ID",
            "Employee Name",
            "Card Number",
            "Transaction Date",
            "Transaction Amount",
            "Transaction Name",
            "Status",
        ]

        for col_idx, expected in enumerate(expected_headers, start=1):
            assert ws.cell(1, col_idx).value == expected

        # Cleanup
        os.remove(file_path)

    def test_only_incomplete_expenses_included(self):
        """Verify only expenses with status != 'Complete' are in Excel."""
        # Create mix of complete and incomplete expenses
        complete_exp = ExpenseTransaction(
            employee_id="EMP001",
            transaction_date=date(2025, 9, 15),
            transaction_amount=Decimal("100.00"),
            transaction_name="Complete",
            has_receipt=True,
            has_gl_code=True,
        )

        incomplete_exp = ExpenseTransaction(
            employee_id="EMP001",
            transaction_date=date(2025, 9, 16),
            transaction_amount=Decimal("50.00"),
            transaction_name="Incomplete",
            has_receipt=False,
            has_gl_code=False,
        )

        employee = Employee(
            employee_id="EMP001",
            name="Test User",
            card_number="1234567890123456",
            expenses=[complete_exp, incomplete_exp],
        )

        session = Session(
            credit_card_pdf_path="test.pdf", expense_report_pdf_path="test2.pdf", employees=[employee]
        )

        file_path = generate_excel_report(session, output_dir="server/tests/unit/temp")

        wb = load_workbook(file_path)
        ws = wb.active

        # Should have header + 1 data row (only incomplete expense)
        assert ws.max_row == 2, "Should have 1 data row (1 incomplete expense)"

        # Verify it's the incomplete one
        assert ws.cell(2, 6).value == "Incomplete"  # Transaction Name column

        os.remove(file_path)

    def test_status_column_values(self):
        """Verify Status column contains correct values."""
        pytest.skip("TODO: Implement status column value tests")

    def test_conditional_formatting(self):
        """Verify conditional formatting applied to Status column."""
        pytest.skip("TODO: Implement conditional formatting tests")
